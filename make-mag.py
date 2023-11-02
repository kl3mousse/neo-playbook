############################################################################
# NEO GEO PLAYBOOK - PDF CREATOR
# https://github.com/kl3mousse/neo-playbook
############################################################################

from includes.get_image import download_image
from includes.mame_command_dat_tools import get_command_blocks
from includes.img_tools import clean_JPG, footer_effect, getImgAspectRatio, crop_bottomright, add_scanlines, crop_upright, getImgSize, image_resize
from includes.hfsdb import get_game_from_hfsdb, hfsdb_scraper, game_data, game_mame_version
from includes.dips import SoftDipsSettings

import datetime, os;

from fpdf import FPDF, HTMLMixin      # to create PDF file
from openpyxl import load_workbook

import json

############################################################################

# constants / prog parameters
NEOGEO_DATA_XLS                = "games.xlsx"
NEOGEO_DATA_XLS_SHEET          = "Games" #switch from "Games" to "Games-test" for testing a smaller chunk of games
NEOGEO_MAME_XLS_SHEET          = "MAME.xml (cleaned)"
COMMAND_DAT_FILE               = "./command-dat/command.dat"
NEOGEO_GAMESGEN                = ["NeoGeo Era", "NeoGeo Resurrection"]  # 'NeoGeo Era' or 'Post NeoGeo' to filter the right set of games 
NEOGEO_MAG_OUTPUT_PDF_PAGESMAX = 90

ct = datetime.datetime.now()
NEOGEO_MAG_OUTPUT_PDF          = "./pdf-output/neo-playbook-alpha-" + str(ct.timestamp())[0:9] # .pdf will be added automatically

# loading secrets from local file
f = open('secrets.json')
secrets=json.load(f)
IGDB_CLIENT_ID=secrets["IGDB_CLIENT_ID"]
IGDB_APP_ACCESS_TOKEN=secrets["IGDB_APP_ACCESS_TOKEN"]
f.close()

############################################################################

############################################################################
def add_cover_page(pdf):
    pdf.add_page()

    pdf.set_xy(0, 70)
    pdf.image("./img-cache/temp-page1.png", x = None, y = None, w = 210, h = 0, type = '', link = '')
    pdf.set_xy(0, 15)
    pdf.image("./img/cover-title-1.png", x = None, y = None, w = 210, h = 0, type = '', link = '')

def add_credits(pdf):
    pdf.add_page()

    pdf.image("./img/neo-playbook-credits.png", x = 0, y = 0, w = 210, h = 297, type = '', link = '')
    
    pdf.set_font("ErbosDracoNova", size = 20)
    pdf.set_text_color(12, 23, 34)
    pdf.set_xy(0, 12)
    pdf.cell(w = 210, h = 12, txt = 'INSERT COIN',  ln = 0, align = 'C')
    
    # load text from HTML file
    pdf.set_font("Osaka", size = 10)
    f = open('./html/credits.html', 'r')
    html_text = f.read()
    f.close()
    pdf.set_xy(40, 45)
    pdf.set_left_margin(35)
    pdf.set_right_margin(35)
    #pdf.write_html(html_text)

def set_page_background(pdf, game, page_num):
    pdf.set_left_margin(5)
    # fill color
    pdf.set_fill_color(123, 134, 145)
    pdf.rect(0, 0, 210, 297, 'F')

    if (page_num % 2) == 0:
        #even / right page
        MAIN_POSX = 13+5
        BAR_POSX = 175
        FOOTER_POSX = 5+5
    else:
        #odd / left page
        MAIN_POSX = 40-5
        BAR_POSX = 0
        FOOTER_POSX = 35-5

    #game background in footer
    if game.game_background is not None:
        pic = download_image(game.game_background)
        bg_h = 40
        bg_w = 160
        pic = clean_JPG(pic)
        crop_upright(pic, bg_w / bg_h, vshift = game.vshift)
        add_scanlines(pic)
        footer_effect(pic, (123, 134, 145))
        pdf.image(pic, x = FOOTER_POSX, y = 297 - bg_h, w = bg_w, h = bg_h, type = '', link = '')

    # page background
    if (page_num % 2) == 0:
        #even page
        pdf.set_xy(0, 0)
        pdf.image("./img/margin-left-2-darkbluesc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
        pdf.set_xy(133, 0)
        pdf.image("./img/margin-right-2-redsc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
    else:
        #odd page
        pdf.set_xy(0, 0)
        pdf.image("./img/margin-left-2-redsc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
        pdf.set_xy(161, 0)
        pdf.image("./img/margin-right-2-darkbluesc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')

    # page number in footer
    if page_num is not None:
        pdf.set_text_color(255, 240, 240)
        pdf.set_font("ErbosDracoNova", size = 8)
        pdf.set_xy(BAR_POSX + 5, 290)
        pdf.cell(w = 25, h = 6, txt = str(page_num),  ln = 0, align = 'C')

def add_moveslist_page(pdf, command_files, game, page_num):
    def add_title():
        pdf.set_y(11)
        pdf.set_x(5)
        pdf.set_text_color(12, 23, 34)
        pdf.set_font("ErbosDracoNova", size = 10)
        page_title = game.title
        pdf.cell(w = 200, h = 9, txt = page_title,  ln = 1, align = 'C')
    
    pdf.add_page()
    added_pages = 1

    set_page_background(pdf, game, page_num + added_pages)

    pdf.set_font("ErbosDracoNova", size = 8)
    pdf.set_text_color(12, 23, 34)
    add_title()
    
    TOP_MARGIN = 20
    BLOCKS_MARGIN = 3
    BLOCK_WIDTH = 60 # was 90
    PAGE_H = 292
    if ((page_num + added_pages) % 2) == 0:
        #even / right page
        LEFT_MARGIN = 20
    else:
        #odd / left page
        LEFT_MARGIN = 6

    current_column = 0 # 0 or 1
    col0_y = TOP_MARGIN
    col1_y = TOP_MARGIN
    col2_y = TOP_MARGIN

    # for each image to add
    for command_file in command_files:
        im_w, im_h = getImgSize(command_file)

        # check if too tall and add new page if required
        if (im_h / im_w * BLOCK_WIDTH) + min(col0_y, col1_y, col2_y) > PAGE_H:
            pdf.add_page()
            added_pages+=1
            if ((page_num + added_pages) % 2) == 0:
                #even / right page
                LEFT_MARGIN = 20
            else:
                #odd / left page
                LEFT_MARGIN = 6
            set_page_background(pdf, game, page_num + added_pages)
            add_title()
            current_column = 0 # 0 or 1
            col0_y = TOP_MARGIN
            col1_y = TOP_MARGIN
            col2_y = TOP_MARGIN

        # add image in column and switch column
        filename = command_file
        im_x = LEFT_MARGIN + current_column * (BLOCK_WIDTH + BLOCKS_MARGIN)
        if current_column == 0: im_y = col0_y
        if current_column == 1: im_y = col1_y
        if current_column == 2: im_y = col2_y
        im_y += BLOCKS_MARGIN
        pdf.image(filename, x = im_x, y = im_y, w = BLOCK_WIDTH, h = round(BLOCK_WIDTH * im_h / im_w), type = '', link = '')
        
        # increment colunms y index
        if current_column == 0: col0_y += round(BLOCK_WIDTH * im_h / im_w) + BLOCKS_MARGIN
        if current_column == 1: col1_y += round(BLOCK_WIDTH * im_h / im_w) + BLOCKS_MARGIN
        if current_column == 2: col2_y += round(BLOCK_WIDTH * im_h / im_w) + BLOCKS_MARGIN
        #if col0_y > col1_y: current_column = 1
        #if col0_y < col1_y: current_column = 0
        if col2_y <= min(col0_y, col1_y):  current_column = 2
        if col1_y <= min(col0_y, col2_y):  current_column = 1
        if col0_y <= min(col1_y, col2_y): current_column = 0


    return added_pages



############################################################################
def add_game_page(pdf, game, page_num):
    pdf.add_page()
    pdf.set_left_margin(5)
    # fill color
    pdf.set_fill_color(123, 134, 145)
    pdf.rect(0, 0, 210, 297, 'F')

    if (page_num % 2) == 0:
        #even / right page
        MAIN_POSX = 13+5
        BAR_POSX = 175
        FOOTER_POSX = 5+5
    else:
        #odd / left page
        MAIN_POSX = 40-5
        BAR_POSX = 0
        FOOTER_POSX = 35-5

    #game background in footer
    if game.game_background is not None:
        pic = download_image(game.game_background)
        bg_h = 40
        bg_w = 160
        pic = clean_JPG(pic)
        crop_upright(pic, bg_w / bg_h, vshift = game.vshift)
        add_scanlines(pic)
        footer_effect(pic, (123, 134, 145))
        pdf.image(pic, x = FOOTER_POSX, y = 297 - bg_h, w = bg_w, h = bg_h, type = '', link = '')

    # page background
    if (page_num % 2) == 0:
        #even page
        pdf.set_xy(0, 0)
        pdf.image("./img/margin-left-2-darkbluesc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
        pdf.set_xy(133, 0)
        pdf.image("./img/margin-right-2-redsc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
    else:
        #odd page
        pdf.set_xy(0, 0)
        pdf.image("./img/margin-left-2-redsc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')
        pdf.set_xy(161, 0)
        pdf.image("./img/margin-right-2-darkbluesc19.png", x = None, y = None, w = 0, h = 297, type = '', link = '')

    # page number in footer
    if page_num is not None:
        pdf.set_text_color(255, 240, 240)
        pdf.set_font("ErbosDracoNova", size = 8)
        pdf.set_xy(BAR_POSX + 5, 290)
        pdf.cell(w = 25, h = 6, txt = str(page_num),  ln = 0, align = 'C')

    # Label for title
    pdf.set_xy(5, 8)
    pdf.set_text_color(20, 20, 20)
    pdf.image("./img/mvs-empty-label-03.png", x = None, y = None, w = 200, h = 0, type = '', link = '')

    # Game title
    pdf.set_y(13)
    pdf.set_font("Osaka", size = 20)
    pdf.cell(w = 200, h = 9, txt = game.title,  ln = 1, align = 'C')

    # Game title in Japanese
    if game.alt_title is None: game.alt_title = " "
    pdf.set_font("Osaka", size = 14)
    pdf.cell(200, 4, txt = game.alt_title, ln = 2, align = 'C')

    # Year & Publisher
    pdf.set_font("Osaka", size = 8)
    pdf.cell(w = 200, h = 3, txt = str(game.year) + ' ' + str(game.publisher) + '            ', ln = 2, align = 'R')

    # NGH/NGM ID
    if game.ngm_id is not None:
        pdf.set_font("Osaka", size = 8)
        pdf.set_xy(21, 26)
        if game.ngm_id < 10: ngm_id_txt = "00" + str(game.ngm_id)
        else:
            if game.ngm_id < 100: ngm_id_txt = "0" + str(game.ngm_id)
            else: ngm_id_txt = str(game.ngm_id)
        
        pdf.cell(w = 10, h = 3, txt = "NGM-" + ngm_id_txt, ln = 0, align = 'L')

    pdf.set_line_width(0.5)
    pdf.set_draw_color(0,0,0)

    #wallpaper
    WALLPAPER_X = MAIN_POSX + 30
    WALLPAPER_Y = 35
    WALLPAPER_H = 70
    WALLPAPER_W = 127
    if game.wallpaper is not None:
        pic = download_image(game.wallpaper)
        crop_bottomright(pic, WALLPAPER_W / WALLPAPER_H)
        pdf.image(pic, x = WALLPAPER_X, y = WALLPAPER_Y, w = WALLPAPER_W, h = WALLPAPER_H, type = '', link = '')
        # draw image border
        pdf.rect(WALLPAPER_X, WALLPAPER_Y, w= WALLPAPER_W, h = WALLPAPER_H)

    #3Dcover
    pdf.set_xy(MAIN_POSX, 40)
    if game.cover3d is not None:
        pic = download_image(game.cover3d)
        if(game.cover3d[-3:].upper() == "JPG"): clean_JPG(pic)
        if(game.cover3d[-4:].upper() == "WEBP"): pic = clean_JPG(pic)
        
        pdf.image(pic, x = None, y = None, w = 35, h = 0, type = '', link = '')

    #screenshots
    SCREENSHOT2_X = MAIN_POSX      #large
    SCREENSHOT2_Y = 107
    SCREENSHOT2_H = 78
    SCREENSHOT2_W = round(SCREENSHOT2_H *320 / 240)

    SCREENSHOT1_X = SCREENSHOT2_X + SCREENSHOT2_W + 2     #title
    SCREENSHOT1_Y = SCREENSHOT2_Y
    SCREENSHOT1_H = 38
    SCREENSHOT1_W = round(SCREENSHOT1_H * 320 / 240)

    SCREENSHOT3_X = SCREENSHOT1_X     #small
    SCREENSHOT3_Y = SCREENSHOT1_Y + SCREENSHOT2_H - SCREENSHOT1_H
    SCREENSHOT3_W = SCREENSHOT1_W
    SCREENSHOT3_H = SCREENSHOT1_H
     
    if (game.invert_ingamescreenshots == "X") or (game.invert_ingamescreenshots == "x"):
        #when X is set in XLS file, then invert the 2 screenshots to get a better choice of main one
        s_temp = game.screenshot2
        game.screenshot2 = game.screenshot3
        game.screenshot3 = s_temp

    if game.screenshot2 is not None:
        pic = download_image(game.screenshot2)
        crop_bottomright(pic, SCREENSHOT2_W / SCREENSHOT2_H)
        pic = clean_JPG(pic)
        add_scanlines(pic)
        pdf.image(pic, x = SCREENSHOT2_X, y = SCREENSHOT2_Y, w = SCREENSHOT2_W, h = SCREENSHOT2_H, type = '', link = '')
        # draw image border
        pdf.rect(SCREENSHOT2_X, SCREENSHOT2_Y, w= SCREENSHOT2_W, h = SCREENSHOT2_H)
    if game.screenshot1 is not None:
        pic = download_image(game.screenshot1)
        crop_bottomright(pic, SCREENSHOT1_W / SCREENSHOT1_H)
        pic = clean_JPG(pic)
        add_scanlines(pic)
        pdf.image(pic, x = SCREENSHOT1_X, y = SCREENSHOT1_Y, w = SCREENSHOT1_W, h = SCREENSHOT1_H, type = '', link = '')
        # draw image border
        pdf.rect(SCREENSHOT1_X, SCREENSHOT1_Y, w= SCREENSHOT1_W, h = SCREENSHOT1_H)
    if game.screenshot3 is not None:
        pic = download_image(game.screenshot3)
        crop_bottomright(pic, SCREENSHOT3_W / SCREENSHOT3_H)
        pic = clean_JPG(pic)
        add_scanlines(pic)
        pdf.image(pic, x = SCREENSHOT3_X, y = SCREENSHOT3_Y, w = SCREENSHOT3_W, h = SCREENSHOT3_H, type = '', link = '')
        # draw image border
        pdf.rect(SCREENSHOT3_X, SCREENSHOT3_Y, w= SCREENSHOT3_W, h = SCREENSHOT3_H)

    # mini marquee
    MINI_MARQUEE_X = SCREENSHOT1_X
    MINI_MARQUEE_Y = SCREENSHOT3_Y + SCREENSHOT3_H + (SCREENSHOT3_Y - (SCREENSHOT1_Y + SCREENSHOT3_H))
    MINI_MARQUEE_W = SCREENSHOT1_W
    MINI_MARQUEE_H = MINI_MARQUEE_W / 11 * 13.75 # mini marquee ratio

    if game.mini_marquee is not None:
        pic = download_image(game.mini_marquee)
        crop_bottomright(pic, MINI_MARQUEE_W / MINI_MARQUEE_H)
        pdf.image(pic, x = MINI_MARQUEE_X, y = MINI_MARQUEE_Y, w = MINI_MARQUEE_W, h = MINI_MARQUEE_H, type = '', link = '')
        # draw image border
        pdf.rect(MINI_MARQUEE_X, MINI_MARQUEE_Y, w= MINI_MARQUEE_W, h = MINI_MARQUEE_H)

    # game soft dips options
    if game.softdipsimage is not None:
        pic = clean_JPG(game.softdipsimage)
        pic = image_resize(pic, 640)
        add_scanlines(pic)
        imw, imh = getImgSize(pic)
        SOFTDIPS_X = SCREENSHOT1_X
        SOFTDIPS_Y = MINI_MARQUEE_Y + MINI_MARQUEE_H + 2
        SOFTDIPS_W = SCREENSHOT1_W
        SOFTDIPS_H = SOFTDIPS_W / imw * imh
        
        pdf.image(pic, x = SOFTDIPS_X, y = SOFTDIPS_Y, w = SOFTDIPS_W, type = '', link = '')
        pdf.rect(SOFTDIPS_X, SOFTDIPS_Y, w= SOFTDIPS_W, h = SOFTDIPS_H)

    # horizontal line
    pdf.set_fill_color(47, 61, 73)
    pdf.rect(SCREENSHOT2_X, MINI_MARQUEE_Y, SCREENSHOT1_W, 1, 'F')
    
    # game description
    if game.description is not None:
        pdf.set_font("Osaka", size = 7)
        #pdf.set_xy(6, 61)
        pdf.set_xy(SCREENSHOT2_X, MINI_MARQUEE_Y +3)
        #pdf.set_text_color(255, 240, 240)
        pdf.set_text_color(47, 61, 73)
        pdf.multi_cell(w = SCREENSHOT2_W, h = 3, txt = game.description, align = 'J')

    # table with MAME versions
    pdf.set_fill_color(47, 61, 73)
    pdf.rect(SCREENSHOT2_X, pdf.get_y() + 3, SCREENSHOT2_W, 1, 'F')
    pdf.set_font("ErbosDracoNova", size = 8)
    pdf.set_xy(SCREENSHOT2_X, pdf.get_y() + 5)
    pdf.cell(w = 17, h = 7, txt = 'versions / roms',  ln = 0, align = 'L')
    #nb_rows = len(game.mame_versions)
    row_i = 0
    for version in game.mame_versions:
        pdf.set_xy(SCREENSHOT2_X, pdf.get_y()+4)

        pdf.set_font("ErbosDracoNova", size = 6)
        ROMNAME_W = 17
        ROMYEAR_W = 4
        pdf.cell(w = ROMNAME_W, h = 5, txt = version.mame_game_rom,  ln = 0, align = 'L')
        pdf.set_font("Osaka", size = 6)
        #pdf.cell(w = ROMYEAR_W, h = 7, txt = str(version.mame_game_year),  ln = 0, align = 'L')
        pdf.cell(w = SCREENSHOT2_W - ROMNAME_W , h = 5, txt = version.mame_game_description,  ln = 0, align = 'L')
        #pdf.cell(w = 26, h = 7, txt = version.mame_game_publisher,  ln = 0, align = 'L')
        #pdf.cell(w = 10, h = 7, txt = version.mame_game_serial,  ln = 0, align = 'L')
        #pdf.cell(w = 10, h = 7, txt = version.mame_game_release,  ln = 0, align = 'L')
        #pdf.cell(w = 9, h = 7, txt = version.mame_game_platform,  ln = 0, align = 'L')
        #pdf.cell(w = 8, h = 7, txt = version.mame_game_compatibility,  ln = 0, align = 'L')
        row_i += 1
    
    # MEGs (size of NeoGeo cart)
    if game.megs is not None:
        pdf.image("img/MEGs.png", x = BAR_POSX+5, y = 70, w = 25, h = 0, type = '', link = '')
        pdf.set_font("ErbosDracoNova", size = 20)
        pdf.set_xy(BAR_POSX+6, 71)
        pdf.cell(w = 25, h = 20, txt = str(game.megs),  ln = 0, align = 'C')

    # type of game (homebrew, proto, licenced...)
    if game.type is not None:
        icon_ok = True
        match game.type:
            case 'Licenced'  : gametype_icon = 'img/icons/type-licenced.png'
            case 'Homebrew'  : gametype_icon = 'img/icons/type-homebrew.png'
            case 'Unreleased': gametype_icon = 'img/icons/type-unreleased.png'
            case 'Unlicenced': gametype_icon = 'img/icons/type-unlicenced.png'
            case 'Hack'      : gametype_icon = 'img/icons/type-hack.png'
            case 'intro demo': gametype_icon = 'img/icons/type-intro-demo.png'
            case 'Bootleg'   : gametype_icon = 'img/icons/type-bootleg.png'
            case 'Port'      : gametype_icon = 'img/icons/type-port.png'
            case _: icon_ok = False
        if icon_ok:
            pdf.image(gametype_icon, x = BAR_POSX+5, y = 110, w = 25, h = 0, type = '', link = '')
    
    # game platforms
    if game.platforms is not None:
        filename = 'img/icons/platform-' + game.platforms + '.png'
        pdf.image(filename, x = BAR_POSX+5, y = 140, w = 25, h = 0, type = '', link = '')


    # genre of game (fight, puzzle, ...)
    if game.genre is not None:
        icon_ok = True
        match game.genre:
            case 'Sport'         : gamegenre_icon = 'img/icons/genre-sport.png'
            case 'Divers'        : gamegenre_icon = 'img/icons/genre-misc.png'
            case 'Shoot them up' : gamegenre_icon = 'img/icons/genre-shootthemup.png'
            case 'Shooter'       : gamegenre_icon = 'img/icons/genre-shootthemup.png'
            case 'Combat'        : gamegenre_icon = 'img/icons/genre-combat.png'
            case 'Reflexion'     : gamegenre_icon = 'img/icons/genre-reflexion.png'
            case 'Plate-formes'  : gamegenre_icon = 'img/icons/genre-platformer.png'
            case 'Quiz'          : gamegenre_icon = 'img/icons/genre-quizz.png'
            case 'Puzzle-Game'   : gamegenre_icon = 'img/icons/genre-puzzle.png'
            case 'Run and gun'   : gamegenre_icon = 'img/icons/genre-runngun.png'
            case 'Beat them all' : gamegenre_icon = 'img/icons/genre-beatthemall.png'
            case 'Action'        : gamegenre_icon = 'img/icons/genre-action.png'
            case 'Course'        : gamegenre_icon = 'img/icons/genre-racing.png'
            case 'RPG'           : gamegenre_icon = 'img/icons/genre-rpg.png'
            case _: icon_ok = False
        if icon_ok:
            pdf.image(gamegenre_icon, x = BAR_POSX+5, y = 170, w = 25, h = 0, type = '', link = '')


    # number of players
    if game.nb_players is not None:
        pdf.image('img/icons/placeholder-blue.png', x = BAR_POSX+5, y = 200, w = 25, h = 0, type = '', link = '')        
        pdf.set_text_color(255, 240, 240)
        pdf.set_font("ErbosDracoNova", size = 8)
        pdf.set_xy(BAR_POSX+3, 210)
        
        if game.nb_players == "2P":
            #pdf.set_xy(BAR_POSX+3, 230)
            #pdf.image("img/2players.png", x = None, y = None, w = 30, h = 0, type = '', link = '')
            
            pdf.cell(w = 30, h = 7, txt = "2 PLAYERS",  ln = 0, align = 'C')
        if game.nb_players == "1P":
            #pdf.set_xy(BAR_POSX+3, 230)
            #pdf.image("img/1player.png", x = None, y = None, w = 30, h = 0, type = '', link = '')
            #pdf.set_font("ErbosDracoNova", size = 8)
            #pdf.set_xy(BAR_POSX+3, 202)
            pdf.cell(w = 30, h = 7, txt = "1 PLAYER",  ln = 0, align = 'C')

############################################################################
# Init PDF file
############################################################################

class MyFPDF(FPDF, HTMLMixin):
    pass

pdf = MyFPDF(unit = "mm", format="A4")
#pdf = FPDF(unit = "mm", format="A4")

pdf.compress = True
#pdf.oversized_images = "WARN"
#pdf.oversized_images = "DOWNSCALE"
#pdf.oversized_images_ratio = 2
pdf.set_left_margin(5)
pdf.set_top_margin(0)
pdf.set_auto_page_break(auto=0)
pdf.add_font("Osaka", "", "fonts/osaka.unicode.ttf", uni=True)
pdf.add_font("ErbosDracoNova", "", "fonts/ErbosdracoNovaOpenNbpRegular-yGa5.ttf", uni=True) 

############################################################################
# Load data from Ms Excel file
############################################################################

wb           = load_workbook(NEOGEO_DATA_XLS, data_only=True)
ws           = wb[NEOGEO_DATA_XLS_SHEET]
ws_mame      = wb[NEOGEO_MAME_XLS_SHEET]
sheet_ranges = wb[NEOGEO_DATA_XLS_SHEET]

# loop in XLS file, for each game found
rowNb = 2
page_type = ws.cell(column=1, row=rowNb).value
#game_id = ws.cell(column=3, row=rowNb).value
page_num = -1 #starts at -1 because of cover

while (not (page_type is None)) :
    page_num += 1

    ### check what type of page and proceed
    match page_type:
        case 'cover_1':
            print('# cover page   ## ', end = '')
            add_cover_page(pdf)
            print('')

        case 'blank':
            print('# blank page   ## ')
            pdf.add_page()

        case 'credits':
            print('# credits page ## ')
            add_credits(pdf)

        case 'game':
            print('# game page    ## ', end = '')
            
            # check if game generation filter is OK, else move next
            game_generation = ws.cell(column=12, row=rowNb).value
            if game_generation not in NEOGEO_GAMESGEN:
                print("")
                print("....skipping game #" + game_id)
                page_num -= 1
            
            else:
                # get info from XLS index
                hfsdb_id  = ws.cell(column=10, row=rowNb).value

                #loads data from HFSdb API:
                game = hfsdb_scraper(hfsdb_id)

                #loads additional data from XLS index
                game.id         = ws.cell(column=3, row=rowNb).value
                game.hfsdb_id   = ws.cell(column=10, row=rowNb).value
                game.title      = ws.cell(column=4, row=rowNb).value
                print("scraping game " + game.title + "....", end = '')
                game.alt_title  = ws.cell(column=7, row=rowNb).value
                game.year       = ws.cell(column=5, row=rowNb).value
                game.publisher  = ws.cell(column=6, row=rowNb).value
                #game_wiki_url  = ws.cell(column=8, row=rowNb).value
                game.type       = ws.cell(column=11, row=rowNb).value
                game.generation = ws.cell(column=12, row=rowNb).value
                
                game.invert_ingamescreenshots = ws.cell(column=13, row=rowNb).value
                game.game_background          = ws.cell(column=14, row=rowNb).value
                game.vshift                   = ws.cell(column=15, row=rowNb).value
                game.ngm_id                   = ws.cell(column=16, row=rowNb).value
                game.megs                     = ws.cell(column=17, row=rowNb).value
                game.platforms                = ws.cell(column=18, row=rowNb).value

                # get roms & boots/hacks data from MAME data
                # loop in XLS file, for each game found
                mame_rowNb = 2
                mame_game_id = ws_mame.cell(column=1, row=mame_rowNb).value
                
                while (mame_game_id is not None) :
                    if mame_game_id == game.id:
                        mame_game_rom           = ws_mame.cell(column=2, row=mame_rowNb).value
                        mame_game_description   = ws_mame.cell(column=3, row=mame_rowNb).value
                        mame_game_year          = ws_mame.cell(column=4, row=mame_rowNb).value
                        mame_game_publisher     = ws_mame.cell(column=5, row=mame_rowNb).value
                        mame_game_serial        = ws_mame.cell(column=6, row=mame_rowNb).value
                        mame_game_release       = ws_mame.cell(column=7, row=mame_rowNb).value
                        mame_game_platform      = ws_mame.cell(column=9, row=mame_rowNb).value
                        mame_game_compatibility = ws_mame.cell(column=10, row=mame_rowNb).value
                        mame_game_excludesoftdips = ws_mame.cell(column=12, row=mame_rowNb).value

                        newmameversion = game_mame_version(mame_game_id, mame_game_rom, mame_game_description, mame_game_year, mame_game_publisher, mame_game_serial,mame_game_release, mame_game_platform, mame_game_compatibility, mame_game_excludesoftdips)

                        game.addrom(newmameversion)
                    # move to next row
                    mame_rowNb += 1
                    c = ws_mame.cell(column=1, row=mame_rowNb)
                    mame_game_id = c.value
                print(str(game.nbroms()) + " roms found..", end = '')

                for rom in game.mame_versions:
                    rom_name = rom.mame_game_rom
                    if rom.mame_game_excludesoftdips != "X":
                        dips = SoftDipsSettings("dips.yaml", "debug_dips.yaml")
                        if not dips.game_settings_found(game_code=rom_name, region='US'):
                            dips.enrich_softdip_settings_from_rom(game_id=rom_name, path= f'./rom/{rom_name}.zip',  language='US')
                            dips.enrich_softdip_settings_from_rom(game_id=rom_name, path= f'./rom/{rom_name}.zip',  language='EU')
                        if dips.game_settings_found(game_code=rom_name, region='US'):        
                            game.softdipsimage = dips.generate_settings_image(game_code=rom_name, region="US", path="./img-cache")
                            game.softdipsimage = dips.generate_settings_image(game_code=rom_name, region="EU", path="./img-cache")
                        else:
                            game.softdipsimage = None

                        # game_dips = SoftDipsSettings(rom_name = rom_name,country= "US", file_path = f'./rom/{rom_name}.zip')
                        # if game_dips.loaded:
                            # game_dips.print_settings()
                        #    softdips_image_filename = game_dips.generate_image('./img-cache/soft-dips')
                        #    print(f'softdips found, saved to {softdips_image_filename}', end = '')
                        #    game.softdipsimage = softdips_image_filename

                #add page in PDF
                add_game_page(pdf, game, page_num )

                #check if moves lists are available for that game in command.dat file from MAME
                commands_found = False
                command_files = []
                for rom in game.mame_versions:
                    rom_name = rom.mame_game_rom

                    command_files = get_command_blocks(rom_name, COMMAND_DAT_FILE)
                    if len(command_files)>0:
                        print(".. moves lists found ! in rom: " + rom_name + " (" + str(len(command_files)) + " blocks)" , end = '')
                        commands_found = True
                        break
                
                #if found, then let's add dedicated page(s) for the moves lists
                if commands_found:
                    added_pages = add_moveslist_page(pdf, command_files, game, page_num)
                    page_num += added_pages

                print(".. done")

        case _:
            break    

    
        
    # move to next row
    rowNb += 1
    c = ws.cell(column=3, row=rowNb)
    game_id = c.value
    c = ws.cell(column=1, row=rowNb)
    page_type = c.value

# Create PDF
# save the pdf with name .pdf
filename = NEOGEO_MAG_OUTPUT_PDF + ".pdf"
print("# writing PDF file: " + filename)
pdf.output(filename) 
