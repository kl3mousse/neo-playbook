"""
One-shot migration: reads games.xlsx and writes data/games.json.

This replaces Excel as the source of truth. Run once, then delete.

Usage:
    uv run python migrate_excel.py                  # migrate "Games" sheet
    uv run python migrate_excel.py --sheet Games-test  # migrate test sheet
"""

import json
import os
import sys
from openpyxl import load_workbook

EXCEL_FILE = "games.xlsx"
GAMES_SHEET = "Games"
MAME_SHEET = "MAME.xml (cleaned)"
OUTPUT_FILE = "data/games.json"


def read_mame_versions(wb):
    """Read MAME sheet and group ROM versions by game_id."""
    ws = wb[MAME_SHEET]
    versions = {}  # game_id -> list of rom dicts
    for row in range(2, ws.max_row + 1):
        game_id = ws.cell(column=1, row=row).value
        if game_id is None:
            break
        rom = {
            "rom_name": ws.cell(column=2, row=row).value,
            "description": ws.cell(column=3, row=row).value,
            "year": ws.cell(column=4, row=row).value,
            "publisher": ws.cell(column=5, row=row).value,
            "serial": ws.cell(column=6, row=row).value,
            "release": ws.cell(column=7, row=row).value,
            "platform_tag": ws.cell(column=9, row=row).value,
            "compatibility": ws.cell(column=10, row=row).value,
            "exclude_softdips": ws.cell(column=12, row=row).value == "X",
        }
        versions.setdefault(game_id, []).append(rom)
    return versions


def migrate_row(ws, row, mame_versions):
    """Convert one Excel row into a JSON entry dict."""
    page_type = ws.cell(column=1, row=row).value

    # Non-game pages are minimal
    if page_type != "game":
        return {"page_type": page_type}

    game_id = ws.cell(column=3, row=row).value
    ngm_id = ws.cell(column=16, row=row).value
    megs = ws.cell(column=17, row=row).value
    bg_url = ws.cell(column=14, row=row).value
    vshift = ws.cell(column=15, row=row).value
    invert = ws.cell(column=13, row=row).value
    platforms = ws.cell(column=18, row=row).value

    entry = {
        "page_type": "game",
        "platform": "neogeo",
        "id": game_id,
        "hfsdb_id": ws.cell(column=10, row=row).value,
        "title": ws.cell(column=4, row=row).value,
        "alt_title": ws.cell(column=7, row=row).value,
        "year": ws.cell(column=5, row=row).value,
        "publisher": ws.cell(column=6, row=row).value,
        "type": ws.cell(column=11, row=row).value,
        "generation": ws.cell(column=12, row=row).value,
        "genre": None,
        "nb_players": None,
        "description": None,
        "images": {
            "wallpaper": {"url": None, "local": None},
            "cover3d": {"url": None, "local": None},
            "screenshot_title": {"url": None, "local": None},
            "screenshot_main": {"url": None, "local": None},
            "screenshot_alt": {"url": None, "local": None},
            "mini_marquee": {"url": None, "local": None},
            "background": {
                "url": bg_url,
                "local": None,
            },
        },
        "background_vshift": int(vshift) if vshift is not None else 0,
        "invert_screenshots": invert in ("X", "x"),
        "platforms": platforms,
        "roms": mame_versions.get(game_id, []),
        "platform_specific": {},
        "softdips_image": None,
        "command_blocks": [],
    }

    if ngm_id is not None:
        try:
            entry["platform_specific"]["ngm_id"] = int(ngm_id)
        except (ValueError, TypeError):
            entry["platform_specific"]["ngm_id"] = str(ngm_id)
    if megs is not None:
        try:
            entry["platform_specific"]["megs"] = int(megs)
        except (ValueError, TypeError):
            entry["platform_specific"]["megs"] = str(megs)

    # Coerce year to string for consistency
    if entry["year"] is not None:
        entry["year"] = str(entry["year"])

    return entry


def migrate(sheet_name=GAMES_SHEET):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    mame_versions = read_mame_versions(wb)
    ws = wb[sheet_name]

    entries = []
    for row in range(2, ws.max_row + 1):
        page_type = ws.cell(column=1, row=row).value
        if page_type is None:
            break
        entries.append(migrate_row(ws, row, mame_versions))

    os.makedirs("data", exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2, ensure_ascii=False)

    games = [e for e in entries if e["page_type"] == "game"]
    roms = sum(len(e.get("roms", [])) for e in games)
    print(f"Migrated {len(entries)} entries ({len(games)} games, {roms} ROM versions) → {OUTPUT_FILE}")


if __name__ == "__main__":
    sheet = GAMES_SHEET
    if "--sheet" in sys.argv:
        idx = sys.argv.index("--sheet")
        sheet = sys.argv[idx + 1]
    migrate(sheet)
